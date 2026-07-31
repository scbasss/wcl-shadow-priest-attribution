#!/usr/bin/env python3
"""
Builds the full Excel workbook from the JSON dumps produced by
raidwide_attribution.py. Three sheets:

  1. "<Priest> Analysis"        - the primary report, all on one tab: summary
                                   numbers, a step-by-step attribution
                                   walkthrough, per-fight and per-player
                                   breakdowns, and methodology notes.
  2. "Top 10 Speed Priests"     - the same raid-wide attribution run against
                                   every report listed in raids_config.json's
                                   top10_groups (e.g. top-10-speed clears for
                                   two raid instances), one table per group.
  3. "Priests vs Warlocks"      - for every report in top10_groups, the
                                   detected shadow priest(s) and every warlock
                                   in that raid side by side, sorted by a
                                   shared "final damage contribution" column
                                   (priest = own damage + raid-wide bonus;
                                   warlock = own damage).

All three sheets are pure openpyxl formulas driven off the JSON - re-run
raidwide_attribution.py against fresh reports and re-run this script to get
updated numbers. openpyxl never computes formulas itself, so after building,
open the file once in Excel/LibreOffice and save it (or run it through
LibreOffice headless: `soffice --headless --convert-to xlsx --outdir . output/attribution.xlsx`)
so the formula cells have cached values before sharing it.

Usage:
  python3 build_workbook.py [--data-dir data] [--config raids_config.json] [--output output/attribution.xlsx]
"""
import argparse
import json
import math
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule

FONT = "Arial"
NAVY = "1F2937"
PRIEST_PURPLE = "6C5CE7"
PRIEST_LIGHT = "EDE9FE"
LOCK_LIGHT = "F3E8FF"
LOCK_ACCENT = "7C3AED"
BAND_LIGHT = "F8FAFC"
BAND_WHITE = "FFFFFF"
GOOD_GREEN = "16A34A"
BT_ACCENT = "DC2626"
HYJAL_ACCENT = "059669"
GOLD = "D97706"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name=FONT, size=11, italic=True, color="64748B")
SECTION_FONT = Font(name=FONT, bold=True, size=13, color="FFFFFF")
LABEL_FONT = Font(name=FONT, bold=True, size=11)
NORMAL_FONT = Font(name=FONT, size=11)
DATA_FONT = Font(name=FONT, size=11, color="0000FF")
FORMULA_FONT = Font(name=FONT, size=11, color="000000")
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = "#,##0"
PCT_FMT = "0.00%"
LINK_FONT = Font(name=FONT, size=11, color="0563C1", underline="single")

INSTANCE_COLORS = {"BT": (BT_ACCENT, "FEE2E2"), "Hyjal": (HYJAL_ACCENT, "D1FAE5")}


def load_json(data_dir, code):
    with open(os.path.join(data_dir, f"raidwide_{code}.json")) as f:
        return json.load(f)


def section_header(ws, r, text, fill=NAVY, span=9):
    c = ws.cell(r, 1, text)
    c.font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    for col in range(1, span + 1):
        ws.cell(r, col).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[r].height = 22
    return r + 2


# ============================================================ Sheet 1 =====

def build_priest_sheet(wb, data_dir, config):
    primary = config["primary_report"]
    D = load_json(data_dir, primary["code"])
    priest_name = D["priest_names"][0] if D["priest_names"] else "Priest"
    report_url = f"https://www.warcraftlogs.com/reports/{D['report_code']}"

    ws = wb.active
    ws.title = f"{priest_name} Analysis"[:31]
    ws.sheet_view.showGridLines = False
    for i, w_ in enumerate([46, 20, 20, 20, 18, 12, 22, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_

    row = 1
    ws.cell(row, 1, f"Shadow Priest ({priest_name}) Debuff Damage Attribution").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    ws.cell(row, 1, "Full raid-wide Shadow Weaving + Misery analysis, all on one tab").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 2

    ws.cell(row, 1, "Warcraft Logs report").font = LABEL_FONT
    lc = ws.cell(row, 2, report_url)
    lc.hyperlink = report_url
    lc.font = LINK_FONT
    row += 2

    row = section_header(ws, row, "SUMMARY")

    def kv(r, label, value_or_formula, is_formula=False, fmt=None, note=None, fill=None):
        lc = ws.cell(r, 1, label)
        lc.font = LABEL_FONT
        c = ws.cell(r, 2, value_or_formula)
        c.font = FORMULA_FONT if is_formula else DATA_FONT
        if fmt:
            c.number_format = fmt
        if note:
            nc = ws.cell(r, 3, note)
            nc.font = NOTE_FONT
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        if fill:
            lc.fill = PatternFill("solid", fgColor=fill)
            c.fill = PatternFill("solid", fgColor=fill)
        return r + 1

    row = kv(row, "Total raid damage (all schools)", D["grand_all_damage"], fmt=NUM_FMT,
             note=f"Sum of every DamageDone event, all sources, all {D.get('num_fights', len(D['fights']))} fights (kills + trash).")
    row_all_damage = row - 1
    row = kv(row, "Total magic-school damage", D["grand_actual_magic"], fmt=NUM_FMT,
             note="Excludes physical (melee/ranged) damage - only schools Misery/Shadow Weaving affect.")
    row_actual_magic = row - 1
    row = kv(row, "Magic damage as % of total raid damage", f"=B{row_actual_magic}/B{row_all_damage}", is_formula=True, fmt=PCT_FMT)
    row = kv(row, "Estimated baseline (no Shadow Weaving/Misery)", D["grand_baseline_magic"], fmt=NUM_FMT)
    row_baseline = row - 1
    row = kv(row, f"TOTAL raid damage attributable to {priest_name}'s debuffs", f"=B{row_actual_magic}-B{row_baseline}", is_formula=True, fmt=NUM_FMT, fill=PRIEST_LIGHT)
    row_total_bonus = row - 1
    row = kv(row, "  ...as % of all magic damage", f"=B{row_total_bonus}/B{row_actual_magic}", is_formula=True, fmt=PCT_FMT, fill=PRIEST_LIGHT)
    row += 1

    row_priest_actual_placeholder = row
    row += 1
    row_priest_bonus_placeholder = row
    row += 1
    row_bonus_to_others_placeholder = row
    row += 1
    row_combined_placeholder = row
    row += 2

    ws.cell(row, 1, "Methodology assumptions (full detail in the Methodology section below):").font = LABEL_FONT
    row += 1
    for note in [
        "Shadow Weaving (ability 15258): +2% shadow damage taken per stack, up to 5 stacks (10% max), 15s duration - stacks read directly from real aura-stack events in the log.",
        "Misery (ability 33200): flat +5% ALL-magic-school damage taken while active, does not stack with itself.",
        "Both multipliers applied multiplicatively per hit, based on whichever stacks were active on that exact target at that exact timestamp.",
        "Physical-school damage excluded entirely since neither debuff affects it.",
        "Pet/guardian/totem damage is rolled up under its owning player (via WCL's petOwner field), so summons like a priest's Shadowfiend or a warlock's demon count toward their owner's total.",
    ]:
        ws.cell(row, 1, "  - " + note).font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
    row += 1

    # ---- attribution walk-through ----
    row = section_header(ws, row, "FULL ATTRIBUTION MATH - STEP BY STEP", fill=LOCK_ACCENT)
    ws.cell(row, 1, f"Every step of how {priest_name}'s combined-impact number above is actually derived, each value linked live to the Summary cells").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 2

    step_hdr_row = row
    for i, h in enumerate(["Step", "What it represents", "Value"], start=1):
        c = ws.cell(step_hdr_row, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[step_hdr_row].height = 22
    row += 1

    steps = [
        ("1", "Total magic-school damage dealt by the WHOLE raid, all sources", f"=B{row_actual_magic}", NUM_FMT, False),
        ("2", "Estimated baseline for that same damage if the tracked debuffs had never been up", f"=B{row_baseline}", NUM_FMT, False),
        ("3", "Step 1 minus Step 2 -> total raid damage attributable to the two debuffs", f"=B{row_actual_magic}-B{row_baseline}", NUM_FMT, True),
        ("4", "Their own personal damage, as logged (this alone is a normal warlock-comparable parse number)", f"=B{row_priest_actual_placeholder}", NUM_FMT, False),
        ("5", "Their own share of Step 3 (the self-buff portion - the debuffs also boost their own hits)", f"=B{row_priest_bonus_placeholder}", NUM_FMT, False),
        ("6", "Step 3 minus Step 5 -> bonus damage the debuffs handed to EVERYONE ELSE in the raid", f"=B{row_priest_bonus_placeholder}", None, False),
    ]
    step_start = row
    for step, desc, formula, fmt, highlight in steps:
        ws.cell(row, 1, step).font = LABEL_FONT
        dc = ws.cell(row, 2, desc)
        dc.font = NORMAL_FONT
        dc.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        vc = ws.cell(row, 8, formula)
        vc.font = FORMULA_FONT
        vc.number_format = fmt or NUM_FMT
        if highlight:
            for col in range(1, 9):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=PRIEST_LIGHT)
        for col in range(1, 9):
            ws.cell(row, col).border = BORDER
        ws.row_dimensions[row].height = 30
        row += 1
    step6_row, step3_row, step5_row = step_start + 5, step_start + 2, step_start + 4
    ws.cell(step6_row, 8, f"=H{step3_row}-H{step5_row}")
    ws.cell(step6_row, 8).number_format = NUM_FMT

    row += 1
    ws.cell(row, 1, "7").font = Font(name=FONT, bold=True, size=12)
    dc = ws.cell(row, 2, "Step 4 plus Step 6 -> their FULL contribution to the raid: own damage + every hit their debuffs made possible for others")
    dc.font = Font(name=FONT, bold=True, size=11)
    dc.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    vc = ws.cell(row, 8, f"=H{step_start+3}+H{step6_row}")
    vc.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
    vc.number_format = NUM_FMT
    for col in range(1, 9):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=LOCK_ACCENT)
        ws.cell(row, col).border = BORDER
    ws.row_dimensions[row].height = 34
    step7_row = row
    row += 2

    ws.cell(row, 1, "For comparison - this raid's warlocks:").font = LABEL_FONT
    row += 1
    for i, h in enumerate(["Player", "Role", "Damage (Buffed)"], start=1):
        c = ws.cell(row, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    row += 1
    ws.cell(row, 1, priest_name).font = Font(name=FONT, bold=True, size=11)
    ws.cell(row, 2, "Shadow Priest (Step 7 total)").font = NORMAL_FONT
    qc = ws.cell(row, 3, f"=H{step7_row}")
    qc.font = FORMULA_FONT
    qc.number_format = NUM_FMT
    for col in range(1, 4):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=PRIEST_LIGHT)
        ws.cell(row, col).border = BORDER
    row += 1

    sources_by_name = {s["name"]: s for s in D["sources"]}
    for lock_name in primary.get("compare_warlocks", []):
        lock = sources_by_name.get(lock_name)
        if not lock:
            continue
        ws.cell(row, 1, lock_name).font = Font(name=FONT, bold=True, size=11)
        ws.cell(row, 2, "Warlock (own damage)").font = NORMAL_FONT
        wc = ws.cell(row, 3, lock["actual"])
        wc.font = DATA_FONT
        wc.number_format = NUM_FMT
        for col in range(1, 4):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=LOCK_LIGHT)
            ws.cell(row, col).border = BORDER
        row += 1
    row += 1

    ws.cell(row, 1, "Why this is the fair comparison number:").font = LABEL_FONT
    row += 1
    for note in [
        "A warlock's parse only ever counts damage THEY personally dealt. Step 4 (own damage) is the equivalent apples-to-apples number, and on its own it's a normal, comparable parse.",
        "But the priest's debuffs also inflate every other caster's magic damage on the same targets - that's Step 6, damage that would NOT exist without them applying Shadow Weaving/Misery.",
        "Step 7 adds those together: not a bigger 'damage meter' number in the traditional sense, but the priest's total measurable effect on the raid's magic damage output, direct and enabled.",
        "See the 'Priests vs Warlocks' tab for how this Step 7 total stacks up against individual warlocks' own damage, across every report in raids_config.json.",
    ]:
        ws.cell(row, 1, "  - " + note).font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
    row += 1

    # ---- Per-Fight ----
    row = section_header(ws, row, "PER-FIGHT BREAKDOWN", fill=NAVY)
    hdr2 = row
    for i, h in enumerate(["Fight ID", "Encounter", "Kill/Trash", "Report Link", "Actual Magic Damage",
                            "Baseline (no debuffs)", "Debuff Bonus", "Bonus %", "All-Schools Total Damage", "Magic % of Total"], start=1):
        c = ws.cell(hdr2, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[hdr2].height = 30
    row += 1

    fight_start = row
    for idx, fight in enumerate(D["fights"]):
        band = BAND_LIGHT if idx % 2 == 0 else BAND_WHITE
        link = f"{report_url}#fight={fight['id']}"
        is_kill = fight.get("kill")
        ws.cell(row, 1, fight["id"]).font = DATA_FONT
        ws.cell(row, 2, fight["name"]).font = DATA_FONT
        kc = ws.cell(row, 3, "Kill" if is_kill else "Trash/Wipe")
        kc.font = Font(name=FONT, size=11, bold=True, color=GOOD_GREEN if is_kill else "94A3B8")
        lc = ws.cell(row, 4, "View fight")
        lc.hyperlink = link
        lc.font = LINK_FONT
        ws.cell(row, 5, fight["actual"]).font = DATA_FONT
        ws.cell(row, 5).number_format = NUM_FMT
        ws.cell(row, 6, f"=E{row}-G{row}").font = FORMULA_FONT
        ws.cell(row, 6).number_format = NUM_FMT
        ws.cell(row, 7, fight["bonus"]).font = DATA_FONT
        ws.cell(row, 7).number_format = NUM_FMT
        ws.cell(row, 8, f"=G{row}/E{row}").font = FORMULA_FONT
        ws.cell(row, 8).number_format = PCT_FMT
        ws.cell(row, 9, fight["all_damage"]).font = DATA_FONT
        ws.cell(row, 9).number_format = NUM_FMT
        ws.cell(row, 10, f"=E{row}/I{row}").font = FORMULA_FONT
        ws.cell(row, 10).number_format = PCT_FMT
        for col in range(1, 11):
            cell = ws.cell(row, col)
            cell.border = BORDER
            if col != 3:
                cell.fill = PatternFill("solid", fgColor=band)
        row += 1
    fight_end = row - 1

    totals_row_fight = row
    ws.cell(totals_row_fight, 2, "TOTALS").font = LABEL_FONT
    for col, formula in [(5, f"=SUM(E{fight_start}:E{fight_end})"), (6, f"=SUM(F{fight_start}:F{fight_end})"),
                          (7, f"=SUM(G{fight_start}:G{fight_end})"), (9, f"=SUM(I{fight_start}:I{fight_end})")]:
        c = ws.cell(totals_row_fight, col, formula)
        c.font = Font(name=FONT, bold=True, color="FFFFFF")
        c.number_format = NUM_FMT
    ws.cell(totals_row_fight, 8, f"=G{totals_row_fight}/E{totals_row_fight}").font = Font(name=FONT, bold=True, color="FFFFFF")
    ws.cell(totals_row_fight, 8).number_format = PCT_FMT
    ws.cell(totals_row_fight, 10, f"=E{totals_row_fight}/I{totals_row_fight}").font = Font(name=FONT, bold=True, color="FFFFFF")
    ws.cell(totals_row_fight, 10).number_format = PCT_FMT
    for col in range(1, 11):
        ws.cell(totals_row_fight, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(totals_row_fight, col).border = BORDER
    row += 2

    # ---- Per-Player ----
    row = section_header(ws, row, "PER-PLAYER BREAKDOWN", fill=NAVY)
    hdr3 = row
    for i, h in enumerate(["Actor ID", "Player/Pet Name", "Actual Magic Damage", "Debuff Bonus", "Baseline (no debuffs)", "Bonus %", "% of Raid's Total Bonus"], start=1):
        c = ws.cell(hdr3, i, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[hdr3].height = 30
    row += 1

    highlight_names = {priest_name} | set(primary.get("compare_warlocks", []))
    player_start = row
    priest_data_row = None
    for idx, src in enumerate(D["sources"]):
        band = BAND_LIGHT if idx % 2 == 0 else BAND_WHITE
        highlight = src["name"] in highlight_names
        fill_color = PRIEST_LIGHT if src["name"] == priest_name else (LOCK_LIGHT if highlight else band)
        if src["name"] == priest_name:
            priest_data_row = row
        ws.cell(row, 1, src["id"]).font = DATA_FONT
        nf = Font(name=FONT, size=11, bold=highlight, color=("0000FF" if not highlight else "1F2937"))
        ws.cell(row, 2, src["name"]).font = nf
        ws.cell(row, 3, src["actual"]).font = DATA_FONT
        ws.cell(row, 3).number_format = NUM_FMT
        ws.cell(row, 4, src["bonus"]).font = DATA_FONT
        ws.cell(row, 4).number_format = NUM_FMT
        ws.cell(row, 5, f"=C{row}-D{row}").font = FORMULA_FONT
        ws.cell(row, 5).number_format = NUM_FMT
        ws.cell(row, 6, f"=D{row}/C{row}").font = FORMULA_FONT
        ws.cell(row, 6).number_format = PCT_FMT
        ws.cell(row, 7, f"=D{row}/$B${row_total_bonus}").font = FORMULA_FONT
        ws.cell(row, 7).number_format = PCT_FMT
        for col in range(1, 8):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=fill_color)
            ws.cell(row, col).border = BORDER
        row += 1
    player_end = row - 1

    totals_row_player = row
    ws.cell(totals_row_player, 2, "TOTALS").font = Font(name=FONT, bold=True, color="FFFFFF")
    for col, formula in [(3, f"=SUM(C{player_start}:C{player_end})"), (4, f"=SUM(D{player_start}:D{player_end})"), (5, f"=SUM(E{player_start}:E{player_end})")]:
        c = ws.cell(totals_row_player, col, formula)
        c.font = Font(name=FONT, bold=True, color="FFFFFF")
        c.number_format = NUM_FMT
    ws.cell(totals_row_player, 6, f"=D{totals_row_player}/C{totals_row_player}").font = Font(name=FONT, bold=True, color="FFFFFF")
    ws.cell(totals_row_player, 6).number_format = PCT_FMT
    for col in range(1, 8):
        ws.cell(totals_row_player, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(totals_row_player, col).border = BORDER
    row += 2

    # backfill Summary placeholders now that we know the priest's per-player row
    ws.cell(row_priest_actual_placeholder, 1, f"{priest_name}'s own damage (self-buffed included)").font = LABEL_FONT
    ws.cell(row_priest_actual_placeholder, 2, f"=C{priest_data_row}").font = FORMULA_FONT
    ws.cell(row_priest_actual_placeholder, 2).number_format = NUM_FMT
    ws.cell(row_priest_bonus_placeholder, 1, f"{priest_name}'s own debuff-bonus (self-buff portion)").font = LABEL_FONT
    ws.cell(row_priest_bonus_placeholder, 2, f"=D{priest_data_row}").font = FORMULA_FONT
    ws.cell(row_priest_bonus_placeholder, 2).number_format = NUM_FMT
    ws.cell(row_bonus_to_others_placeholder, 1, "Bonus damage given to the REST of the raid").font = LABEL_FONT
    ws.cell(row_bonus_to_others_placeholder, 2, f"=B{row_total_bonus}-B{row_priest_bonus_placeholder}").font = FORMULA_FONT
    ws.cell(row_bonus_to_others_placeholder, 2).number_format = NUM_FMT
    ws.cell(row_combined_placeholder, 1, f"{priest_name} combined impact (own damage + enabled in others)").font = LABEL_FONT
    ws.cell(row_combined_placeholder, 2, f"=B{row_priest_actual_placeholder}+B{row_bonus_to_others_placeholder}").font = FORMULA_FONT
    ws.cell(row_combined_placeholder, 2).number_format = NUM_FMT
    ws.cell(row_combined_placeholder, 3, "NOT a normal damage-meter number - personal output plus the extra damage the debuffs let everyone else do.").font = NOTE_FONT
    ws.merge_cells(start_row=row_combined_placeholder, start_column=3, end_row=row_combined_placeholder, end_column=9)
    for rr in (row_priest_actual_placeholder, row_priest_bonus_placeholder, row_bonus_to_others_placeholder, row_combined_placeholder):
        fill = PRIEST_LIGHT if rr == row_combined_placeholder else "FFFFFF"
        ws.cell(rr, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(rr, 2).fill = PatternFill("solid", fgColor=fill)

    # ---- Methodology ----
    row = section_header(ws, row, "METHODOLOGY, SOURCES, AND KNOWN LIMITATIONS", fill=NAVY)
    lines = [
        "DATA SOURCE",
        f"Warcraft Logs v2 GraphQL API ({report_url}), pulled via the OAuth 'authorization_code' user flow "
        "(client_credentials alone can't read an archived report - that needs a personal access token tied to a subscribing account).",
        "",
        "DEBUFFS TRACKED",
        "- Shadow Weaving (ability ID 15258): Shadow Priest talent proc. +2% shadow damage taken per stack, up to 5 stacks (10% max), 15s duration per application. Source: wowhead.com/tbc/spell=15334/shadow-weaving",
        "  Note: WCL's own UI displays this aura's debuff table entry as 'Shadow Vulnerability' - same underlying ability ID (15258), just a different display label.",
        "- Misery (ability ID 33200): +5% ALL-magic-school damage taken while active. Flat, does NOT stack with itself. Source: wowhead.com/tbc/spell=33195/misery",
        "",
        "CALCULATION",
        "For every magic-school damage event in the raid: look up whichever of the two debuffs (and Shadow Weaving's exact stack count) were active on that specific target at that exact timestamp, "
        "using the real aura-application/refresh/removal events pulled from the log. Multiply the two effects together to get the total multiplier active on that hit. "
        "Baseline damage = actual damage / multiplier. The difference (actual - baseline) is the portion of that hit attributable to the debuffs.",
        "",
        "PHYSICAL DAMAGE EXCLUSION",
        "Neither debuff affects physical damage, so every damage event's ability was checked against the report's own ability-type catalog and physical-school hits (type '1') were excluded entirely.",
        "",
        "DEBUFF EVENT VISIBILITY",
        "The events query defaults to hostilityType: Friendlies for the 'Debuffs' data type; hostilityType: Enemies had to be passed explicitly since these debuffs only ever land on hostile targets.",
        "",
        "PET/GUARDIAN ROLLUP",
        "Pets, guardians, and totems are separate actors in WCL's data. Their damage is rolled up under the owning player via the petOwner field, so summons (a priest's Shadowfiend, a warlock's demon) "
        "count toward that player's own total instead of disappearing into an untracked bucket.",
        "",
        "SCOPE",
        f"All {D.get('num_fights', len(D['fights']))} fights in this report were attempted - kills and trash/wipe pulls alike. "
        f"{len(D['fights'])} completed successfully" + (f" ({D.get('num_fights', len(D['fights'])) - len(D['fights'])} dropped due to transient API errors)." if D.get('num_fights', len(D['fights'])) > len(D['fights']) else "."),
    ]
    for line in lines:
        is_header = line.isupper() and line != ""
        cell = ws.cell(row, 1, line)
        cell.font = LABEL_FONT if is_header else NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        if is_header:
            cell.fill = PatternFill("solid", fgColor="EEF2FF")
            for col in range(1, 10):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="EEF2FF")
        else:
            n_lines = max(1, math.ceil(len(line) / 140))
            ws.row_dimensions[row].height = max(15, n_lines * 14)
        row += 1

    ws.freeze_panes = "A6"
    ws.sheet_properties.tabColor = PRIEST_PURPLE


# ============================================================ Sheet 2 =====

def build_top10_sheet(wb, data_dir, config):
    ws = wb.create_sheet("Top 10 Speed Priests"[:31])
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD
    for i, w_ in enumerate([6, 24, 12, 12, 16, 10, 20, 18, 18, 18, 10, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_

    row = 1
    ws.cell(row, 1, "Shadow Priest Debuff Impact - Top 10 Speed Clears").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 1
    ws.cell(row, 1, "Same raid-wide attribution as the primary analysis tab, run against every report in raids_config.json's top10_groups").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 2

    ws.cell(row, 1, "Zone rankings source").font = LABEL_FONT
    zone_url = config.get("zone_rankings_url", "")
    lc = ws.cell(row, 2, zone_url)
    lc.hyperlink = zone_url
    lc.font = LINK_FONT
    row += 2

    headers = ["Rank", "Guild", "Realm", "Clear Time", "Date", "Report", "Shadow Priest",
               "Total Magic Dmg", "Baseline (no debuffs)", "Raid-wide Debuff Bonus", "Bonus %",
               "Priest's Own Dmg", "Priest's Own Bonus"]
    incomplete_notes = []
    group_ranges = []

    for group in config.get("top10_groups", []):
        accent, light = INSTANCE_COLORS.get(group["instance"], (NAVY, BAND_LIGHT))
        row = section_header(ws, row, group["title"], fill=accent, span=13)
        hdr_row = row
        for i, h in enumerate(headers, start=1):
            c = ws.cell(hdr_row, i, h)
            c.font = HEADER_FONT
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            c.border = BORDER
        ws.row_dimensions[hdr_row].height = 30
        row += 1
        data_start = row
        for idx, raid in enumerate(group["raids"]):
            d = load_json(data_dir, raid["code"])
            priest_name = ", ".join(d["priest_names"]) if d["priest_names"] else "(none detected)"
            priest_actual = sum(p["actual"] for p in d["priests"])
            priest_bonus = sum(p["bonus"] for p in d["priests"])
            band = light if idx % 2 == 0 else BAND_WHITE
            report_url = f"https://www.warcraftlogs.com/reports/{raid['code']}"

            ws.cell(row, 1, raid["rank"]).font = DATA_FONT
            ws.cell(row, 2, raid["guild"]).font = Font(name=FONT, bold=True, size=11)
            ws.cell(row, 3, raid["realm"]).font = NORMAL_FONT
            ws.cell(row, 4, raid["time"]).font = NORMAL_FONT
            ws.cell(row, 5, raid["date"]).font = NORMAL_FONT
            lc = ws.cell(row, 6, "View report")
            lc.hyperlink = report_url
            lc.font = LINK_FONT
            ws.cell(row, 7, priest_name).font = Font(name=FONT, bold=True, size=11)
            ws.cell(row, 8, d["grand_actual_magic"]).font = DATA_FONT
            ws.cell(row, 8).number_format = NUM_FMT
            ws.cell(row, 9, f"=H{row}-J{row}").font = FORMULA_FONT
            ws.cell(row, 9).number_format = NUM_FMT
            ws.cell(row, 10, d["grand_bonus"]).font = DATA_FONT
            ws.cell(row, 10).number_format = NUM_FMT
            ws.cell(row, 11, f"=J{row}/H{row}").font = FORMULA_FONT
            ws.cell(row, 11).number_format = PCT_FMT
            ws.cell(row, 12, priest_actual).font = DATA_FONT
            ws.cell(row, 12).number_format = NUM_FMT
            ws.cell(row, 13, priest_bonus).font = DATA_FONT
            ws.cell(row, 13).number_format = NUM_FMT
            for col in range(1, 14):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=band)
                ws.cell(row, col).border = BORDER
            expected = d.get("num_fights", len(d["fights"]))
            if expected > len(d["fights"]):
                incomplete_notes.append(f"  - {d['label']} ({raid['code']}): {expected - len(d['fights'])} of {expected} fights failed with a transient API error and are excluded from this row's totals.")
            row += 1
        data_end = row - 1
        group_ranges.append((data_start, data_end))

        totals_row = row
        ws.cell(totals_row, 2, "AVERAGE / TOTAL").font = Font(name=FONT, bold=True, color="FFFFFF")
        for col, formula in [(8, f"=SUM(H{data_start}:H{data_end})"), (9, f"=SUM(I{data_start}:I{data_end})"),
                              (10, f"=SUM(J{data_start}:J{data_end})"), (12, f"=SUM(L{data_start}:L{data_end})"),
                              (13, f"=SUM(M{data_start}:M{data_end})")]:
            c = ws.cell(totals_row, col, formula)
            c.font = Font(name=FONT, bold=True, color="FFFFFF")
            c.number_format = NUM_FMT
        ws.cell(totals_row, 11, f"=AVERAGE(K{data_start}:K{data_end})").font = Font(name=FONT, bold=True, color="FFFFFF")
        ws.cell(totals_row, 11).number_format = PCT_FMT
        for col in range(1, 14):
            ws.cell(totals_row, col).fill = PatternFill("solid", fgColor=accent)
            ws.cell(totals_row, col).border = BORDER
        row = totals_row + 2

    if incomplete_notes:
        ws.cell(row, 1, "Note on data completeness:").font = LABEL_FONT
        row += 1
        for note in incomplete_notes:
            ws.cell(row, 1, note).font = NOTE_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            row += 1
        row += 1

    if group_ranges:
        row = section_header(ws, row, "BONUS % ACROSS ALL REPORTS (CHART SOURCE DATA)", fill=GOLD, span=13)
        chart_hdr = row
        for i, h in enumerate(["Report", "Instance", "Bonus %"], start=1):
            ws.cell(row, i, h).font = LABEL_FONT
        row += 1
        chart_first = row
        for group, (data_start, data_end) in zip(config.get("top10_groups", []), group_ranges):
            for rr in range(data_start, data_end + 1):
                ws.cell(row, 1, f"=B{rr}&\" (#\"&A{rr}&\")\"")
                ws.cell(row, 2, group["instance"])
                ws.cell(row, 3, f"=K{rr}")
                row += 1
        chart_last = row - 1

        chart = BarChart()
        chart.type = "bar"
        chart.title = "Debuff Bonus % of Magic Damage - All Reports"
        chart.y_axis.title = "Report"
        chart.x_axis.title = "Bonus %"
        chart.style = 10
        chart.height = 16
        chart.width = 24
        cats = Reference(ws, min_col=1, min_row=chart_first, max_row=chart_last)
        data = Reference(ws, min_col=3, min_row=chart_hdr, max_row=chart_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = GOLD
        ws.add_chart(chart, f"E{chart_hdr}")

    ws.freeze_panes = "A6"


# ============================================================ Sheet 3 =====

def build_comparison_sheet(wb, data_dir, config):
    ws = wb.create_sheet("Priests vs Warlocks"[:31])
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PRIEST_PURPLE
    for i, w_ in enumerate([22, 8, 20, 14, 20, 20, 20, 10, 26, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w_

    row = 1
    ws.cell(row, 1, "Shadow Priest vs. Warlocks - Every Top-Speed Raid").font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    ws.cell(row, 1, "Each raid's shadow priest next to every warlock in that raid, sorted by a shared final-contribution number").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    headers = ["Raid", "Rank", "Player", "Role", "Damage (Buffed, as logged)",
               "Damage (Unbuffed by Priest)", "Bonus from Priest", "Bonus %",
               "Final Damage Contribution", ""]

    for group in config.get("top10_groups", []):
        accent, _light = INSTANCE_COLORS.get(group["instance"], (NAVY, BAND_LIGHT))
        c = ws.cell(row, 1, group["title"].upper())
        c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        for col in range(1, 11):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.row_dimensions[row].height = 20
        row += 2

        for raid in group["raids"]:
            d = load_json(data_dir, raid["code"])
            report_url = f"https://www.warcraftlogs.com/reports/{raid['code']}"
            grand_bonus = d["grand_bonus"]

            label = f"{group['instance']} #{raid['rank']} - {raid['guild']}"
            c = ws.cell(row, 1, label)
            c.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            lc = ws.cell(row, 10, "Report")
            lc.hyperlink = report_url
            lc.font = Font(name=FONT, size=9, color="FFFFFF", underline="single")
            for col in range(1, 11):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=accent)
            ws.row_dimensions[row].height = 18
            row += 1

            for i, h in enumerate(headers[:9], start=1):
                hc = ws.cell(row, i, h)
                hc.font = Font(name=FONT, bold=True, size=9, color="475569")
                hc.alignment = Alignment(wrap_text=True, vertical="center")
                hc.border = BORDER
            ws.row_dimensions[row].height = 26
            row += 1

            combined = []
            for pr in d["priests"]:
                combined.append({
                    "name": pr["name"], "role": "Shadow Priest",
                    "actual": pr["actual"], "bonus": pr["bonus"],
                    "final": pr["actual"] + grand_bonus,
                    "final_formula_tpl": "=E{r}+" + str(grand_bonus),
                })
            for lock in d["warlocks"]:
                combined.append({
                    "name": lock["name"], "role": "Warlock",
                    "actual": lock["actual"], "bonus": lock["bonus"],
                    "final": lock["actual"],
                    "final_formula_tpl": "=E{r}",
                })
            combined.sort(key=lambda x: -x["final"])

            for entry in combined:
                is_priest = entry["role"] == "Shadow Priest"
                fill_color = "FFFFFF" if is_priest else LOCK_LIGHT
                ws.cell(row, 3, entry["name"]).font = Font(name=FONT, bold=is_priest, size=11)
                ws.cell(row, 4, entry["role"]).font = NORMAL_FONT
                ws.cell(row, 5, entry["actual"]).font = DATA_FONT
                ws.cell(row, 5).number_format = NUM_FMT
                ws.cell(row, 6, f"=E{row}-G{row}").font = FORMULA_FONT
                ws.cell(row, 6).number_format = NUM_FMT
                ws.cell(row, 7, entry["bonus"]).font = DATA_FONT
                ws.cell(row, 7).number_format = NUM_FMT
                ws.cell(row, 8, f"=G{row}/E{row}" if entry["actual"] else 0).font = FORMULA_FONT
                ws.cell(row, 8).number_format = PCT_FMT
                contrib_cell = ws.cell(row, 9, entry["final_formula_tpl"].format(r=row))
                contrib_cell.font = FORMULA_FONT
                contrib_cell.number_format = NUM_FMT
                for col in range(1, 10):
                    ws.cell(row, col).fill = PatternFill("solid", fgColor=fill_color)
                    ws.cell(row, col).border = BORDER
                row += 1
            row += 1  # spacer between raids

    ws.cell(row, 1, "Note:").font = LABEL_FONT
    row += 1
    ws.cell(row, 1, "\"Damage (Unbuffed by Priest)\" = Buffed Damage minus the debuff bonus, i.e. what that player's damage would have been with zero tracked-debuff uptime on their targets. "
                     "It is NOT their damage with the priest removed from the raid entirely - it isolates specifically the tracked-debuff slice.").font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    ws.cell(row, 1, "\"Final Damage Contribution\" = for the priest, their own buffed damage PLUS the full raid-wide debuff damage attributable to them (every source in the raid, not just themself); "
                     "for a warlock, it's simply their own buffed damage. Same column, so the priest's whole measurable effect on the raid sits directly against each warlock's solo parse.").font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    ws.freeze_panes = "A5"


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--data-dir", default="data", help="Directory containing raidwide_<code>.json files (default: ./data)")
    parser.add_argument("--config", default="raids_config.json", help="Path to the raid-list config (default: ./raids_config.json)")
    parser.add_argument("--output", default="output/attribution.xlsx", help="Output .xlsx path (default: ./output/attribution.xlsx)")
    args = parser.parse_args()

    with open(args.config) as f:
        config = json.load(f)

    wb = Workbook()
    build_priest_sheet(wb, args.data_dir, config)
    if config.get("top10_groups"):
        build_top10_sheet(wb, args.data_dir, config)
        build_comparison_sheet(wb, args.data_dir, config)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print("Formulas won't show cached values until this file is opened once in Excel/LibreOffice and saved")
    print("(or run through `soffice --headless --convert-to xlsx --outdir <dir> <file>`).")


if __name__ == "__main__":
    main()
